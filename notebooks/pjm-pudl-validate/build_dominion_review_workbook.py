from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
DOMINION_DIR = BASE_DIR / "csvs" / "Dominion"
OUTPUT_PATH = DOMINION_DIR / "dominion_pudl_review.xlsx"

SHEET_ORDER = [
    ("validation_summary_snapshot.csv", "01_Validation"),
    ("dominion_solar_summary.csv", "02_DOM_Solar"),
    ("dominion_solar_alias_crosswalk.csv", "03_DOM_Solar_Alias"),
    ("dominion_solar_alias_adjusted_comparison.csv", "04_DOM_Solar_Adjust"),
    ("dominion_solar_matches.csv", "05_DOM_Solar_Match"),
    ("dominion_solar_workbook_unmatched_after_alias_crosswalk.csv", "06_DOM_Solar_WB_Un"),
    ("dominion_solar_pudl_missing_after_fuzzy.csv", "07_DOM_Solar_PUDL_Un"),
    ("capacity_by_fuel_comparison.csv", "08_Capacity_Fuel"),
    ("renewable_capacity_comparison.csv", "09_Renewables"),
    ("plant_capacity_matches_best_available.csv", "10_Best_Matches"),
    ("plant_capacity_matches_fuzzy_only.csv", "11_Fuzzy_Only"),
    ("plant_capacity_matches_exact_name.csv", "12_Exact_Matches"),
    ("workbook_plants_unmatched_after_fuzzy.csv", "13_WB_Unmatched"),
    ("pudl_plants_missing_from_workbook_after_fuzzy_ge_100mw.csv", "14_PUDL_Miss_100"),
    ("pudl_plants_missing_from_workbook_after_fuzzy_all.csv", "15_PUDL_Miss_All"),
    ("workbook_plants_aggregated.csv", "16_WB_Plants"),
    ("pudl_plants_aggregated_latest.csv", "17_PUDL_Plants"),
    ("subset_definition.csv", "18_Subset_Def"),
    ("comparison_years.csv", "19_Years"),
    ("workbook_stack_active_units.csv", "20_WB_Units"),
    ("workbook_pjm_raw_data.csv", "21_WB_Raw"),
    ("pudl_operating_generators_latest.csv", "22_PUDL_Gens"),
    ("pudl_operating_generators_heat_rate_year.csv", "23_PUDL_HR_Year"),
    ("heat_rate_compare_exact_name.csv", "24_HR_Compare"),
    ("heat_rate_outliers_gt_20pct.csv", "25_HR_Outliers"),
    ("retired_generators_found_in_workbook.csv", "26_Retired"),
    ("pudl_pjm_plants_reference.csv", "27_PUDL_Plant_Ref"),
    ("csv_export_manifest.csv", "28_Manifest"),
]


def autosize_sheet(worksheet) -> None:
    widths: dict[int, int] = {}
    for row in worksheet.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            text = "" if value is None else str(value)
            widths[idx] = min(max(widths.get(idx, 0), len(text) + 2), 60)
    for idx, width in widths.items():
        worksheet.column_dimensions[get_column_letter(idx)].width = width


def style_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    autosize_sheet(worksheet)


def build_start_here(manifest: pd.DataFrame) -> pd.DataFrame:
    descriptions = manifest.set_index("file_name")["description"].to_dict()
    steps = [
        ("01_Validation", "validation_summary_snapshot.csv", "Read the patched Dominion validation status first."),
        ("02_DOM_Solar", "dominion_solar_summary.csv", "Use this as the high-level Dominion solar summary."),
        ("03_DOM_Solar_Alias", "dominion_solar_alias_crosswalk.csv", "Review the explicit solar alias crosswalk next."),
        ("04_DOM_Solar_Adjust", "dominion_solar_alias_adjusted_comparison.csv", "Then inspect the alias-adjusted one-project-per-row solar comparison."),
        ("05_DOM_Solar_Match", "dominion_solar_matches.csv", "Then review the solar plants that match after exact and fuzzy logic."),
        ("06_DOM_Solar_WB_Un", "dominion_solar_workbook_unmatched_after_alias_crosswalk.csv", "Then inspect the remaining workbook solar plants after aliases are removed."),
        ("07_DOM_Solar_PUDL_Un", "dominion_solar_pudl_missing_after_fuzzy.csv", "Then inspect remaining PUDL solar plants."),
        ("08_Capacity_Fuel", "capacity_by_fuel_comparison.csv", "Top-level fuel bucket comparison."),
        ("10_Best_Matches", "plant_capacity_matches_best_available.csv", "Best plant matches across all fuels."),
        ("13_WB_Unmatched", "workbook_plants_unmatched_after_fuzzy.csv", "Largest workbook plants still unmatched."),
        ("14_PUDL_Miss_100", "pudl_plants_missing_from_workbook_after_fuzzy_ge_100mw.csv", "Largest PUDL plants still unmatched."),
        ("16_WB_Plants", "workbook_plants_aggregated.csv", "Workbook plant rollup for tracing."),
        ("17_PUDL_Plants", "pudl_plants_aggregated_latest.csv", "PUDL plant rollup for tracing."),
        ("20_WB_Units", "workbook_stack_active_units.csv", "Workbook unit-level detail."),
        ("22_PUDL_Gens", "pudl_operating_generators_latest.csv", "PUDL generator-level detail."),
        ("18_Subset_Def", "subset_definition.csv", "Subset definition and fuzzy matching rule."),
        ("19_Years", "comparison_years.csv", "Source years and PJM filter metadata."),
        ("28_Manifest", "csv_export_manifest.csv", "Full inventory of workbook tabs."),
    ]
    rows = []
    for idx, (sheet, csv_name, why) in enumerate(steps, start=1):
        rows.append(
            {
                "step": idx,
                "sheet_name": sheet,
                "source_csv": csv_name,
                "why_start_here": why,
                "what_to_look_for": descriptions.get(csv_name, ""),
            }
        )
    return pd.DataFrame(rows)


def main() -> None:
    manifest_path = DOMINION_DIR / "csv_export_manifest.csv"
    if not manifest_path.exists():
        raise FileNotFoundError(f"Missing manifest: {manifest_path}")

    manifest = pd.read_csv(manifest_path)
    start_here = build_start_here(manifest)

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        start_here.to_excel(writer, sheet_name="00_Start_Here", index=False)
        for csv_name, sheet_name in SHEET_ORDER:
            path = DOMINION_DIR / csv_name
            if not path.exists():
                raise FileNotFoundError(f"Missing CSV: {path}")
            pd.read_csv(path).to_excel(writer, sheet_name=sheet_name, index=False)
        for worksheet in writer.book.worksheets:
            style_sheet(worksheet)

    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
